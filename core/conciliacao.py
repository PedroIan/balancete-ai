"""
core/conciliacao.py
Conciliação bancária e geração do arquivo XLSX final.
Não faz chamadas LLM, não lê arquivos, não conhece Streamlit.
"""

from __future__ import annotations

import copy
import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paleta de cores ──────────────────────────────────────────────────────────
_COR_CABECALHO = "1F3864"     # azul escuro
_COR_RECEITA   = "E2EFDA"     # verde claro
_COR_DESPESA   = "FCE4D6"     # laranja claro
_COR_SUSPEITO  = "FFF2CC"     # amarelo claro

_BORDA_FINA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_FORMATO_MOEDA = 'R$ #,##0.00'
_TOLERANCIA_VALOR = 0.05       # R$ 0,05
_TOLERANCIA_DIAS  = 3          # ±3 dias úteis


# ── API pública ──────────────────────────────────────────────────────────────

def gerar_xlsx(
    transacoes: List[Dict],
    extrato_movs: List[Dict],
    competencia: str,
    saldo_inicial: float,
) -> bytes:
    """Gera XLSX completo do zero. Retorna bytes prontos para download."""
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    ok, suspeitos = separar_suspeitos(transacoes)
    receitas = [t for t in ok if t["tipo"] == "receita"]
    despesas = [t for t in ok if t["tipo"] == "despesa"]
    # suspeitos entram nos totais mas aparecem também na aba Revisar
    receitas_total = sum(t["valor"] for t in ok + suspeitos if t["tipo"] == "receita")
    despesas_total = sum(t["valor"] for t in ok + suspeitos if t["tipo"] == "despesa")
    saldo_final = saldo_inicial + receitas_total - despesas_total

    _criar_aba_resumo(wb, competencia, saldo_inicial, receitas_total, despesas_total, saldo_final, transacoes)

    if receitas:
        _criar_aba_receitas(wb, receitas, competencia)

    despesas_completas = [t for t in transacoes if t["tipo"] == "despesa"]
    if despesas_completas:
        _criar_aba_despesas(wb, despesas_completas, competencia)

    if extrato_movs:
        _criar_aba_extrato(wb, extrato_movs, competencia)

        debitos_extrato = [m for m in extrato_movs if m["tipo"] == "debito"]
        if despesas_completas and debitos_extrato:
            pares, ext_sem, tx_sem = conciliar(extrato_movs, despesas_completas)
            _criar_aba_conciliacao(wb, pares, ext_sem, tx_sem, competencia)

    if suspeitos:
        _criar_aba_revisar(wb, suspeitos, competencia)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def preencher_template(
    template_bytes: bytes,
    transacoes: List[Dict],
    extrato_movs: List[Dict],
    competencia: str,
    saldo_inicial: float,
) -> bytes:
    """Preenche template XLSX do usuário com placeholders e adiciona abas extras."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

    ok, suspeitos = separar_suspeitos(transacoes)
    receitas_total = sum(t["valor"] for t in transacoes if t["tipo"] == "receita")
    despesas_total = sum(t["valor"] for t in transacoes if t["tipo"] == "despesa")
    saldo_final = saldo_inicial + receitas_total - despesas_total

    substituicoes = {
        "{{competencia}}": competencia,
        "{{saldo_inicial}}": f"{saldo_inicial:.2f}",
        "{{total_receitas}}": f"{receitas_total:.2f}",
        "{{total_despesas}}": f"{despesas_total:.2f}",
        "{{saldo_final}}": f"{saldo_final:.2f}",
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "s" and cell.value:
                    for placeholder, valor in substituicoes.items():
                        if placeholder in str(cell.value):
                            cell.value = str(cell.value).replace(placeholder, valor)

    # Adiciona abas extras ao final
    despesas = [t for t in transacoes if t["tipo"] == "despesa"]
    if extrato_movs:
        _criar_aba_extrato(wb, extrato_movs, competencia)
        debitos = [m for m in extrato_movs if m["tipo"] == "debito"]
        if despesas and debitos:
            pares, ext_sem, tx_sem = conciliar(extrato_movs, despesas)
            _criar_aba_conciliacao(wb, pares, ext_sem, tx_sem, competencia)

    if suspeitos:
        _criar_aba_revisar(wb, suspeitos, competencia)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def conciliar(
    extrato_movs: List[Dict],
    transacoes: List[Dict],
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    Best-match entre débitos do extrato e despesas do balancete.
    Critérios: valor ±R$0,05 e data ±3 dias.
    Desempate: menor diferença de data, depois menor diferença de valor.
    Retorna: (pares, extrato_sem_par, transacoes_sem_par)
    """
    debitos = [m for m in extrato_movs if m["tipo"] == "debito"]
    despesas = [t for t in transacoes if t["tipo"] == "despesa"]

    disponivel_debitos = list(range(len(debitos)))
    disponivel_despesas = list(range(len(despesas)))

    # Monta candidatos: (diff_dias, diff_valor, idx_debito, idx_despesa)
    candidatos = []
    for i, deb in enumerate(debitos):
        for j, desp in enumerate(despesas):
            diff_v = abs(deb["valor"] - desp["valor"])
            if diff_v > _TOLERANCIA_VALOR:
                continue
            diff_d = abs(_dias_entre(deb["data"], desp["data"]))
            if diff_d > _TOLERANCIA_DIAS:
                continue
            candidatos.append((diff_d, diff_v, i, j))

    candidatos.sort()

    pares: List[Dict] = []
    matched_deb: set = set()
    matched_desp: set = set()

    for diff_d, diff_v, i, j in candidatos:
        if i in matched_deb or j in matched_desp:
            continue
        matched_deb.add(i)
        matched_desp.add(j)
        pares.append({
            "extrato": debitos[i],
            "balancete": despesas[j],
            "diff_dias": diff_d,
            "diff_valor": diff_v,
        })

    ext_sem = [debitos[i] for i in range(len(debitos)) if i not in matched_deb]
    tx_sem  = [despesas[j] for j in range(len(despesas)) if j not in matched_desp]

    return (pares, ext_sem, tx_sem)


def separar_suspeitos(
    transacoes: List[Dict],
) -> Tuple[List[Dict], List[Dict]]:
    """Separa transações em (ok, suspeitas). Suspeitas não são removidas dos totais."""
    ok = [t for t in transacoes if not t.get("suspeito")]
    suspeitos = [t for t in transacoes if t.get("suspeito")]
    return (ok, suspeitos)


# ── Criação de abas ──────────────────────────────────────────────────────────

def _criar_aba_resumo(
    wb: Workbook,
    competencia: str,
    saldo_inicial: float,
    receitas_total: float,
    despesas_total: float,
    saldo_final: float,
    transacoes: List[Dict],
) -> None:
    ws = wb.create_sheet("Resumo")

    _cabecalho_aba(ws, f"Balancete Condominial — {competencia}", 2)

    linhas = [
        ("Saldo Inicial", saldo_inicial),
        ("(+) Total Receitas", receitas_total),
        ("(-) Total Despesas", despesas_total),
        ("= Saldo Final", saldo_final),
    ]
    for texto, valor in linhas:
        linha = ws.max_row + 1
        ws.cell(linha, 1, texto).font = Font(bold=True)
        ws.cell(linha, 2, valor).number_format = _FORMATO_MOEDA

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def _criar_aba_receitas(
    wb: Workbook, receitas: List[Dict], competencia: str
) -> None:
    ws = wb.create_sheet("Receitas")
    cabecalhos = ["Data", "Histórico", "Unidade/Pagador", "Categoria", "Valor", "Fonte"]
    _linha_cabecalho(ws, cabecalhos)

    for t in sorted(receitas, key=lambda x: x["data"]):
        ws.append([
            t["data"],
            t["descricao"],
            t["fornecedor"],
            t["categoria"],
            t["valor"],
            t["fonte"],
        ])
        ws.cell(ws.max_row, 5).number_format = _FORMATO_MOEDA
        _colorir_linha(ws, ws.max_row, _COR_RECEITA, len(cabecalhos))

    _autofit(ws, cabecalhos)


def _criar_aba_despesas(
    wb: Workbook, despesas: List[Dict], competencia: str
) -> None:
    ws = wb.create_sheet("Despesas")
    cabecalhos = ["Data", "Fornecedor", "CNPJ", "Histórico", "Categoria", "Valor", "Fonte"]
    _linha_cabecalho(ws, cabecalhos)

    for t in sorted(despesas, key=lambda x: x["data"]):
        ws.append([
            t["data"],
            t["fornecedor"],
            t.get("cnpj", ""),
            t["descricao"],
            t["categoria"],
            t["valor"],
            t["fonte"],
        ])
        ws.cell(ws.max_row, 6).number_format = _FORMATO_MOEDA
        cor = _COR_SUSPEITO if t.get("suspeito") else _COR_DESPESA
        _colorir_linha(ws, ws.max_row, cor, len(cabecalhos))

    _autofit(ws, cabecalhos)


def _criar_aba_extrato(
    wb: Workbook, movs: List[Dict], competencia: str
) -> None:
    ws = wb.create_sheet("Extrato (Referência)")
    cabecalhos = ["Data", "Descrição", "Tipo", "Valor", "Saldo", "Fonte"]
    _linha_cabecalho(ws, cabecalhos)

    for m in sorted(movs, key=lambda x: x["data"]):
        ws.append([
            m["data"],
            m["descricao"],
            m["tipo"],
            m["valor"],
            m.get("saldo"),
            m["fonte"],
        ])
        ws.cell(ws.max_row, 4).number_format = _FORMATO_MOEDA
        if m.get("saldo") is not None:
            ws.cell(ws.max_row, 5).number_format = _FORMATO_MOEDA

    _autofit(ws, cabecalhos)


def _criar_aba_conciliacao(
    wb: Workbook,
    pares: List[Dict],
    ext_sem: List[Dict],
    tx_sem: List[Dict],
    competencia: str,
) -> None:
    ws = wb.create_sheet("Conciliação")
    cabecalhos = [
        "Status", "Data Extrato", "Descrição Extrato", "Valor Extrato",
        "Data Balancete", "Fornecedor Balancete", "Valor Balancete", "Δ Dias", "Δ Valor"
    ]
    _linha_cabecalho(ws, cabecalhos)

    for par in pares:
        ext = par["extrato"]
        bal = par["balancete"]
        ws.append([
            "🟢 Conciliado",
            ext["data"], ext["descricao"], ext["valor"],
            bal["data"], bal["fornecedor"], bal["valor"],
            par["diff_dias"], par["diff_valor"],
        ])
        ws.cell(ws.max_row, 4).number_format = _FORMATO_MOEDA
        ws.cell(ws.max_row, 7).number_format = _FORMATO_MOEDA

    for m in ext_sem:
        ws.append([
            "🔴 Só no extrato",
            m["data"], m["descricao"], m["valor"],
            None, None, None, None, None,
        ])
        ws.cell(ws.max_row, 4).number_format = _FORMATO_MOEDA

    for t in tx_sem:
        ws.append([
            "🔵 Só no balancete",
            None, None, None,
            t["data"], t["fornecedor"], t["valor"],
            None, None,
        ])
        ws.cell(ws.max_row, 7).number_format = _FORMATO_MOEDA

    _autofit(ws, cabecalhos)


def _criar_aba_revisar(
    wb: Workbook, suspeitos: List[Dict], competencia: str
) -> None:
    ws = wb.create_sheet("⚠️ Revisar")
    cabecalhos = ["Data", "Fornecedor", "Descrição", "Valor", "Tipo", "Categoria", "Motivo", "Fonte"]
    _linha_cabecalho(ws, cabecalhos)

    for t in suspeitos:
        motivos = []
        if t["valor"] == 0.0:
            motivos.append("valor zero")
        if not t.get("descricao"):
            motivos.append("sem descrição")
        if not t.get("data"):
            motivos.append("sem data")
        if t.get("cnpj") and len(t["cnpj"]) == 14:
            from core.classifier import _validar_cnpj
            _, valido = _validar_cnpj(t["cnpj"])
            if not valido:
                motivos.append("CNPJ inválido")

        ws.append([
            t["data"],
            t["fornecedor"],
            t["descricao"],
            t["valor"],
            t["tipo"],
            t["categoria"],
            "; ".join(motivos) or "suspeito",
            t["fonte"],
        ])
        ws.cell(ws.max_row, 4).number_format = _FORMATO_MOEDA
        _colorir_linha(ws, ws.max_row, _COR_SUSPEITO, len(cabecalhos))

    _autofit(ws, cabecalhos)


# ── Helpers de formatação ────────────────────────────────────────────────────

def _cabecalho_aba(ws, titulo: str, n_colunas: int) -> None:
    ws.merge_cells(f"A1:{get_column_letter(n_colunas)}1")
    cel = ws.cell(1, 1, titulo)
    cel.font = Font(bold=True, color="FFFFFF", size=13)
    cel.fill = PatternFill("solid", fgColor=_COR_CABECALHO)
    cel.alignment = Alignment(horizontal="center")


def _linha_cabecalho(ws, cabecalhos: List[str]) -> None:
    ws.append(cabecalhos)
    linha = ws.max_row
    for col, _ in enumerate(cabecalhos, start=1):
        cel = ws.cell(linha, col)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=_COR_CABECALHO)
        cel.alignment = Alignment(horizontal="center")
        cel.border = _BORDA_FINA


def _colorir_linha(ws, linha: int, cor_hex: str, n_colunas: int) -> None:
    fill = PatternFill("solid", fgColor=cor_hex)
    for col in range(1, n_colunas + 1):
        ws.cell(linha, col).fill = fill
        ws.cell(linha, col).border = _BORDA_FINA


def _autofit(ws, cabecalhos: List[str]) -> None:
    larguras_base = [max(len(str(h)), 10) for h in cabecalhos]
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            if i >= len(larguras_base):
                break
            larguras_base[i] = max(larguras_base[i], len(str(cell.value or "")))
    for i, largura in enumerate(larguras_base, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(largura + 2, 50)


def _dias_entre(data1: str, data2: str) -> int:
    try:
        d1 = date.fromisoformat(data1)
        d2 = date.fromisoformat(data2)
        return (d1 - d2).days
    except Exception:
        return 9999
